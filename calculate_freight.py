#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
快递运费计算脚本 - 基于ExpressFreightCalculator项目的默认报价表
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime
import math
from collections import defaultdict

# ===================== 默认报价表数据 =====================

# 区域-省份映射
REGION_MAPPING = {
    "一区": ["江苏", "浙江", "安徽", "上海", "江苏省", "浙江省", "安徽省", "上海市"],
    "二区": ["山东", "广东", "福建", "北京", "河南", "湖北", "湖南", "江西", "天津", "河北",
            "山东省", "广东省", "福建省", "北京市", "河南省", "湖北省", "湖南省", "江西省", "天津市", "河北省"],
    "三区": ["山西", "广西", "四川", "重庆", "陕西", "贵州", "辽宁", "吉林", "黑龙江", "云南",
            "山西省", "广西省", "四川省", "重庆市", "陕西省", "贵州省", "辽宁省", "吉林省", "黑龙江省", "云南省"],
    "四区": ["海南", "甘肃", "青海", "内蒙古", "宁夏",
            "海南省", "甘肃省", "青海省", "内蒙古", "宁夏省"],
    "五区": ["新疆", "西藏", "新疆省", "西藏省"]
}

# 简化省份名（去掉"省"、"市"）
def normalize_province(prov):
    if prov:
        prov = prov.strip()
        # 去掉省、市、自治区等后缀
        if prov.endswith("省") or prov.endswith("市"):
            prov = prov[:-1]
        elif prov == "内蒙古自治区":
            prov = "内蒙古"
        elif prov == "广西壮族自治区":
            prov = "广西"
        elif prov == "新疆维吾尔自治区" or prov == "新疆维吾尔族自治区":
            prov = "新疆"
        elif prov == "西藏自治区":
            prov = "西藏"
    return prov

# 价格规则（按区域）
PRICE_RULES = {
    "一区": [
        {"min": 0, "max": 0.5, "price": 2.26, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 2.46, "type": "fixed"},
        {"min": 1, "max": 2, "price": 3.56, "type": "fixed"},
        {"min": 2, "max": 3, "price": 4.76, "type": "fixed"},
        {"min": 3, "max": 30, "first": 3.76, "add": 0.8, "type": "standard"},
        {"min": 30, "max": -1, "first": 3.86, "add": 0.8, "type": "standard"},
    ],
    "二区": [
        {"min": 0, "max": 0.5, "price": 2.26, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 2.46, "type": "fixed"},
        {"min": 1, "max": 2, "price": 3.56, "type": "fixed"},
        {"min": 2, "max": 3, "price": 4.76, "type": "fixed"},
        {"min": 3, "max": 30, "first": 3.76, "add": 1.1, "type": "standard"},
        {"min": 30, "max": -1, "first": 4.06, "add": 1.3, "type": "standard"},
    ],
    "三区": [
        {"min": 0, "max": 0.5, "price": 2.26, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 2.46, "type": "fixed"},
        {"min": 1, "max": 2, "price": 3.56, "type": "fixed"},
        {"min": 2, "max": 3, "price": 4.76, "type": "fixed"},
        {"min": 3, "max": 30, "first": 3.76, "add": 1.5, "type": "standard"},
        {"min": 30, "max": -1, "first": 4.06, "add": 1.6, "type": "standard"},
    ],
    "四区": [
        {"min": 0, "max": 0.5, "price": 2.56, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 3.56, "type": "fixed"},
        {"min": 1, "max": 2, "price": 4.06, "type": "fixed"},
        {"min": 2, "max": 3, "price": 5.06, "type": "fixed"},
        {"min": 3, "max": 30, "first": 3.76, "add": 2.5, "type": "standard"},
        {"min": 30, "max": -1, "first": 4.06, "add": 4.3, "type": "standard"},
    ],
    "新疆": [
        {"min": 0, "max": 0.5, "price": 10, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 13, "type": "fixed"},
        {"min": 1, "max": 2, "price": 20, "type": "fixed"},
        {"min": 2, "max": 3, "price": 25, "type": "fixed"},
        {"min": 3, "max": 30, "first": 15, "add": 15, "type": "standard"},
        {"min": 30, "max": -1, "first": 15, "add": 15, "type": "standard"},
    ],
    "西藏": [
        {"min": 0, "max": 0.5, "price": 13, "type": "fixed"},
        {"min": 0.51, "max": 1, "price": 15, "type": "fixed"},
        {"min": 1, "max": 2, "price": 25, "type": "fixed"},
        {"min": 2, "max": 3, "price": 30, "type": "fixed"},
        {"min": 3, "max": 30, "first": 15, "add": 15, "type": "standard"},
        {"min": 30, "max": -1, "first": 15, "add": 15, "type": "standard"},
    ],
}

def find_region(province):
    """根据省份查找区域"""
    prov = normalize_province(province)
    
    # 先检查新疆西藏（特殊价格）
    if prov in ["新疆", "西藏"]:
        return prov
    
    for region, provinces in REGION_MAPPING.items():
        if prov in provinces or prov + "省" in provinces or prov + "市" in provinces:
            return region
    
    # 默认返回一区
    return "一区"

def calculate_freight(weight, province):
    """计算运费"""
    if weight <= 0:
        return 0
    
    region = find_region(province)
    rules = PRICE_RULES.get(region, PRICE_RULES["一区"])
    
    for rule in rules:
        if rule["type"] == "fixed":
            # 固定价格段
            if weight >= rule["min"] and (rule["max"] == -1 or weight < rule["max"]):
                return rule["price"]
        else:
            # 标准计算：首重 + 续重
            if weight >= rule["min"] and (rule["max"] == -1 or weight < rule["max"]):
                if weight <= rule["min"]:
                    return rule["first"]
                extra_weight = weight - rule["min"]
                extra_kg = math.ceil(extra_weight)
                return rule["first"] + extra_kg * rule["add"]
    
    # 未匹配到规则
    return 0

def process_excel(input_file, output_file):
    """处理Excel文件"""
    print(f"正在读取: {input_file}")
    
    wb_in = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws_in = wb_in.active
    
    # 创建输出文件
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "运费计算结果"
    
    # 写入表头
    headers = ["业务时间", "运单号", "结算重量(kg)", "目的省份", "体积重(kg)", 
               "订单客户", "客户", "计算重量(kg)", "运费(元)", "区域"]
    for col, h in enumerate(headers, 1):
        ws_out.cell(row=1, column=col, value=h)
    
    # 统计数据
    total_count = 0
    error_count = 0
    total_freight = 0
    region_stats = defaultdict(lambda: {"count": 0, "freight": 0})
    
    # 读取原始表头获取列索引
    header_row = next(ws_in.iter_rows(min_row=1, max_row=1))
    col_map = {}
    for cell in header_row:
        col_map[cell.value] = cell.column
    
    print("列映射:", col_map)
    print("开始处理数据...")
    
    # 处理数据行
    row_idx = 2
    for row in ws_in.iter_rows(min_row=2):
        try:
            # 读取各列数据
            business_time = row[col_map["业务时间"] - 1].value if "业务时间" in col_map else None
            waybill_no = row[col_map["运单号"] - 1].value if "运单号" in col_map else None
            weight = row[col_map["结算重量"] - 1].value if "结算重量" in col_map else 0
            province = row[col_map["目的省份"] - 1].value if "目的省份" in col_map else ""
            volumetric = row[col_map["体积重"] - 1].value if "体积重" in col_map else 0
            order_customer = row[col_map["订单客户"] - 1].value if "订单客户" in col_map else ""
            customer = row[col_map["客户"] - 1].value if "客户" in col_map else ""
            
            # 计算重量（使用结算重量，体积重为空）
            calc_weight = weight if weight else 0
            
            # 计算运费
            freight = calculate_freight(calc_weight, province)
            
            # 查找区域
            region = find_region(province)
            
            # 写入结果
            ws_out.cell(row=row_idx, column=1, value=business_time)
            ws_out.cell(row=row_idx, column=2, value=waybill_no)
            ws_out.cell(row=row_idx, column=3, value=weight)
            ws_out.cell(row=row_idx, column=4, value=province)
            ws_out.cell(row=row_idx, column=5, value=volumetric)
            ws_out.cell(row=row_idx, column=6, value=order_customer)
            ws_out.cell(row=row_idx, column=7, value=customer)
            ws_out.cell(row=row_idx, column=8, value=calc_weight)
            ws_out.cell(row=row_idx, column=9, value=freight)
            ws_out.cell(row=row_idx, column=10, value=region)
            
            # 统计
            total_count += 1
            total_freight += freight
            region_stats[region]["count"] += 1
            region_stats[region]["freight"] += freight
            
            if freight == 0 and calc_weight > 0:
                error_count += 1
            
            # 进度显示
            if total_count % 10000 == 0:
                print(f"已处理 {total_count} 条，运费合计: {total_freight:.2f} 元")
            
            row_idx += 1
            
        except Exception as e:
            error_count += 1
            if error_count % 100 == 0:
                print(f"错误: {e}")
    
    wb_in.close()
    
    # 写入汇总信息
    summary_row = row_idx + 2
    ws_out.cell(row=summary_row, column=1, value="===== 汇总统计 =====")
    ws_out.cell(row=summary_row + 1, column=1, value="总条数")
    ws_out.cell(row=summary_row + 1, column=2, value=total_count)
    ws_out.cell(row=summary_row + 2, column=1, value="总运费(元)")
    ws_out.cell(row=summary_row + 2, column=2, value=round(total_freight, 2))
    ws_out.cell(row=summary_row + 3, column=1, value="错误条数")
    ws_out.cell(row=summary_row + 3, column=2, value=error_count)
    
    # 各区域统计
    ws_out.cell(row=summary_row + 5, column=1, value="===== 各区域统计 =====")
    for i, (region, stats) in enumerate(sorted(region_stats.items(), key=lambda x: x[1]["freight"], reverse=True)):
        ws_out.cell(row=summary_row + 6 + i, column=1, value=region)
        ws_out.cell(row=summary_row + 6 + i, column=2, value=stats["count"])
        ws_out.cell(row=summary_row + 6 + i, column=3, value=stats["freight"])
        ws_out.cell(row=summary_row + 6 + i, column=4, value=round(stats["freight"] / stats["count"], 2) if stats["count"] > 0 else 0)
    
    # 保存
    print(f"\n正在保存结果到: {output_file}")
    wb_out.save(output_file)
    
    print(f"\n===== 处理完成 =====")
    print(f"总条数: {total_count}")
    print(f"总运费: {total_freight:.2f} 元")
    print(f"错误条数: {error_count}")
    print(f"\n各区域统计:")
    for region, stats in sorted(region_stats.items(), key=lambda x: x[1]["freight"], reverse=True):
        avg = stats["freight"] / stats["count"] if stats["count"] > 0 else 0
        print(f"  {region}: {stats['count']}条, 总运费 {stats['freight']:.2f}元, 平均 {avg:.2f}元/件")

if __name__ == "__main__":
    input_file = "蜜丝婷-4月发件账单表1.xlsx"
    output_file = "蜜丝婷-4月运费计算结果.xlsx"
    process_excel(input_file, output_file)